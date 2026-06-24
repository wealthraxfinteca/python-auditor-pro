"""
Report Generator
Produces styled Excel and PDF audit reports
"""

import pandas as pd
import numpy as np
import io
from datetime import datetime
from typing import Dict, Optional
import logging

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataLabels
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Generates styled Excel audit reports with embedded charts
    and a PDF summary.
    """

    # Excel style definitions
    STYLES = {
        'header_fill': PatternFill(
            start_color='1F4E79',
            end_color='1F4E79',
            fill_type='solid'
        ),
        'subheader_fill': PatternFill(
            start_color='2E75B6',
            end_color='2E75B6',
            fill_type='solid'
        ),
        'high_fill': PatternFill(
            start_color='FF0000',
            end_color='FF0000',
            fill_type='solid'
        ),
        'medium_fill': PatternFill(
            start_color='FFC000',
            end_color='FFC000',
            fill_type='solid'
        ),
        'low_fill': PatternFill(
            start_color='00B0F0',
            end_color='00B0F0',
            fill_type='solid'
        ),
        'pass_fill': PatternFill(
            start_color='00B050',
            end_color='00B050',
            fill_type='solid'
        ),
        'fail_fill': PatternFill(
            start_color='FF0000',
            end_color='FF0000',
            fill_type='solid'
        ),
        'alt_row_fill': PatternFill(
            start_color='DCE6F1',
            end_color='DCE6F1',
            fill_type='solid'
        ),
        'white_font': Font(
            name='Calibri', bold=True, color='FFFFFF', size=11
        ),
        'header_font': Font(
            name='Calibri', bold=True, color='FFFFFF', size=12
        ),
        'title_font': Font(
            name='Calibri', bold=True, color='1F4E79', size=16
        ),
        'bold_font': Font(name='Calibri', bold=True, size=11),
        'normal_font': Font(name='Calibri', size=10),
        'center_align': Alignment(
            horizontal='center', vertical='center', wrap_text=True
        ),
        'left_align': Alignment(
            horizontal='left', vertical='center'
        ),
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

    def generate_excel_report(
        self,
        findings: Dict,
        report_name: str = "Audit_Report"
    ) -> bytes:
        """
        Generate a fully styled Excel audit report.
        Returns the report as bytes for Streamlit download.
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Cover Page
        self._create_cover_sheet(wb, findings, report_name)

        # Sheet 2: Audit Summary
        self._create_summary_sheet(wb, findings)

        # Sheet 3: Detailed Findings
        if not findings.get('flagged_findings', pd.DataFrame()).empty:
            self._create_findings_sheet(
                wb, findings['flagged_findings']
            )

        # Sheet 4: Value Changes
        if not findings.get('value_differences', pd.DataFrame()).empty:
            self._create_changes_sheet(
                wb, findings['value_differences']
            )

        # Sheet 5: Added Records
        added = findings.get('added_records', pd.DataFrame())
        if not added.empty:
            self._create_data_sheet(
                wb, added, 'Added Records',
                'Records present in File 2 only'
            )

        # Sheet 6: Removed Records
        removed = findings.get('removed_records', pd.DataFrame())
        if not removed.empty:
            self._create_data_sheet(
                wb, removed, 'Removed Records',
                'Records present in File 1 only'
            )

        # Sheet 7: Reconciliation Stats
        self._create_recon_sheet(wb, findings)

        # Save to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_multi_file_report(
        self,
        multi_findings: Dict,
        report_name: str = "Multi_Audit_Report"
    ) -> bytes:
        """Generate report for multi-file comparison."""
        wb = Workbook()
        wb.remove(wb.active)

        # Cover
        self._create_multi_cover_sheet(wb, multi_findings, report_name)

        # Reconciliation Matrix
        recon_matrix = multi_findings.get('reconciliation_matrix')
        if recon_matrix is not None and not recon_matrix.empty:
            self._create_data_sheet(
                wb,
                recon_matrix,
                'Recon Matrix',
                'Cross-File Reconciliation Summary'
            )

        # Individual comparisons
        for comp_label, comp_data in multi_findings.get(
            'comparisons', {}
        ).items():
            short_label = comp_label[:25]

            findings_df = comp_data.get(
                'flagged_findings', pd.DataFrame()
            )
            if not findings_df.empty:
                sheet_name = f"Findings_{short_label}"[:31]
                self._create_findings_sheet(
                    wb, findings_df, sheet_name
                )

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _create_cover_sheet(
        self,
        wb: Workbook,
        findings: Dict,
        report_name: str
    ):
        """Create a styled cover/title page."""
        ws = wb.create_sheet("Cover Page")
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30

        # Title block
        ws.merge_cells('B2:C2')
        title_cell = ws['B2']
        title_cell.value = "🔍 AUDIT COMPARISON REPORT"
        title_cell.font = Font(
            name='Calibri', bold=True, size=22, color='1F4E79'
        )
        title_cell.alignment = self.STYLES['center_align']
        ws.row_dimensions[2].height = 40

        ws.merge_cells('B3:C3')
        subtitle = ws['B3']
        subtitle.value = report_name.replace('_', ' ')
        subtitle.font = Font(
            name='Calibri', size=14, color='2E75B6', italic=True
        )
        subtitle.alignment = self.STYLES['center_align']

        # Info block
        stats = findings.get('recon_stats', {})
        metrics = findings.get('variance_metrics', {})

        info_rows = [
            ("Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Comparison:", findings.get('comparison_label', 'N/A')),
            ("Primary Key:", findings.get('primary_key', 'N/A')),
            ("Match Rate:", f"{stats.get('match_rate_pct', 0):.1f}%"),
            ("Audit Status:", (
                "✅ PASS"
                if stats.get('audit_pass', False)
                else "❌ FAIL"
            )),
            ("Total Differences:", str(
                metrics.get('total_differences', 0)
            )),
            ("Records Added:", str(stats.get('records_added', 0))),
            ("Records Removed:", str(stats.get('records_removed', 0)))
        ]

        for i, (label, value) in enumerate(info_rows, start=5):
            row = i + 1
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = self.STYLES['bold_font']
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = self.STYLES['normal_font']

            if 'PASS' in str(value):
                ws[f'C{row}'].font = Font(
                    name='Calibri', color='00B050',
                    bold=True, size=11
                )
            elif 'FAIL' in str(value):
                ws[f'C{row}'].font = Font(
                    name='Calibri', color='FF0000',
                    bold=True, size=11
                )

        ws.sheet_view.showGridLines = False

    def _create_summary_sheet(self, wb: Workbook, findings: Dict):
        """Create audit summary metrics sheet."""
        ws = wb.create_sheet("Audit Summary")
        stats = findings.get('recon_stats', {})
        metrics = findings.get('variance_metrics', {})

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "AUDIT SUMMARY METRICS"
        ws['A1'].font = self.STYLES['title_font']
        ws['A1'].alignment = self.STYLES['center_align']
        ws.row_dimensions[1].height = 32

        # Headers
        headers = ['Audit Metric', 'Value', 'Status']
        self._write_header_row(ws, 3, headers)

        # Data rows
        audit_pass = stats.get('audit_pass', False)
        summary_data = [
            (
                "Total Records (File 1)",
                list(stats.values())[0] if stats else 0,
                'INFO'
            ),
            (
                "Total Records (File 2)",
                list(stats.values())[1] if len(stats) > 1 else 0,
                'INFO'
            ),
            (
                "Common Records Analyzed",
                stats.get('common_records', 0),
                'INFO'
            ),
            (
                "Clean Matches",
                stats.get('matched_clean', 0),
                'PASS'
            ),
            (
                "Records with Changes",
                stats.get('records_with_changes', 0),
                'FAIL' if stats.get('records_with_changes', 0) > 0 else 'PASS'
            ),
            (
                "Records Added (File 2 only)",
                stats.get('records_added', 0),
                'WARN' if stats.get('records_added', 0) > 0 else 'PASS'
            ),
            (
                "Records Removed (File 1 only)",
                stats.get('records_removed', 0),
                'WARN' if stats.get('records_removed', 0) > 0 else 'PASS'
            ),
            (
                "Total Value Differences",
                metrics.get('total_differences', 0),
                'FAIL' if metrics.get('total_differences', 0) > 0 else 'PASS'
            ),
            (
                "Numeric Differences",
                metrics.get('numeric_differences', 0),
                'INFO'
            ),
            (
                "Text Differences",
                metrics.get('text_differences', 0),
                'INFO'
            ),
            (
                "Total Numeric Variance",
                f"{metrics.get('total_variance', 0):,.2f}",
                'FAIL' if abs(metrics.get('total_variance', 0)) > 0 else 'PASS'
            ),
            (
                "Match Rate %",
                f"{stats.get('match_rate_pct', 0):.2f}%",
                'PASS' if audit_pass else 'FAIL'
            ),
            (
                "Audit Result",
                'PASSED' if audit_pass else 'FAILED',
                'PASS' if audit_pass else 'FAIL'
            )
        ]

        for i, (metric, value, status) in enumerate(summary_data, start=4):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'C{i}'] = status

            ws[f'A{i}'].font = self.STYLES['normal_font']
            ws[f'B{i}'].font = self.STYLES['bold_font']
            ws[f'B{i}'].alignment = Alignment(horizontal='center')

            # Status color coding
            status_fills = {
                'PASS': self.STYLES['pass_fill'],
                'FAIL': self.STYLES['fail_fill'],
                'WARN': self.STYLES['medium_fill'],
                'INFO': self.STYLES['subheader_fill']
            }
            if status in status_fills:
                ws[f'C{i}'].fill = status_fills[status]
                ws[f'C{i}'].font = self.STYLES['white_font']
                ws[f'C{i}'].alignment = self.STYLES['center_align']

            if i % 2 == 0:
                ws[f'A{i}'].fill = self.STYLES['alt_row_fill']
                ws[f'B{i}'].fill = self.STYLES['alt_row_fill']

        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 14
        ws.sheet_view.showGridLines = False

    def _create_findings_sheet(
        self,
        wb: Workbook,
        flagged_df: pd.DataFrame,
        sheet_name: str = "Flagged Findings"
    ):
        """Create the detailed flagged findings sheet."""
        ws = wb.create_sheet(sheet_name)

        ws.merge_cells('A1:H1')
        ws['A1'] = "FLAGGED DISCREPANCIES — DETAILED AUDIT FINDINGS"
        ws['A1'].font = self.STYLES['title_font']
        ws['A1'].alignment = self.STYLES['center_align']
        ws.row_dimensions[1].height = 30

        # Write headers
        headers = list(flagged_df.columns)
        self._write_header_row(ws, 3, headers)

        # Write data with severity coloring
        severity_fills = {
            'HIGH': self.STYLES['high_fill'],
            'MEDIUM': self.STYLES['medium_fill'],
            'LOW': self.STYLES['low_fill']
        }

        for row_idx, (_, row) in enumerate(
            flagged_df.iterrows(), start=4
        ):
            severity = row.get('Severity', 'LOW')
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.STYLES['normal_font']
                cell.border = self.STYLES['thin_border']

                if col_idx == len(headers):  # Severity column
                    if severity in severity_fills:
                        cell.fill = severity_fills[severity]
                        cell.font = self.STYLES['white_font']
                        cell.alignment = self.STYLES['center_align']

                if row_idx % 2 == 0 and col_idx != len(headers):
                    cell.fill = self.STYLES['alt_row_fill']

        # Auto-width columns
        for col_idx, col in enumerate(flagged_df.columns, start=1):
            max_len = max(
                len(str(col)),
                flagged_df[col].astype(str).str.len().max()
            )
            ws.column_dimensions[
                get_column_letter(col_idx)
            ].width = min(max_len + 4, 40)

        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A4'

    def _create_changes_sheet(
        self,
        wb: Workbook,
        diff_df: pd.DataFrame
    ):
        """Create value changes detail sheet."""
        ws = wb.create_sheet("Value Changes")
        ws.merge_cells('A1:G1')
        ws['A1'] = "DETAILED VALUE CHANGES"
        ws['A1'].font = self.STYLES['title_font']
        ws['A1'].alignment = self.STYLES['center_align']
        ws.row_dimensions[1].height = 28

        headers = list(diff_df.columns)
        self._write_header_row(ws, 3, headers)

        for row_idx, (_, row) in enumerate(diff_df.iterrows(), start=4):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if pd.isna(value):
                    cell.value = ''
                elif isinstance(value, float):
                    cell.value = round(value, 4)
                else:
                    cell.value = value

                cell.font = self.STYLES['normal_font']
                if row_idx % 2 == 0:
                    cell.fill = self.STYLES['alt_row_fill']

        for col_idx, col in enumerate(diff_df.columns, start=1):
            ws.column_dimensions[
                get_column_letter(col_idx)
            ].width = 22

        ws.freeze_panes = 'A4'
        ws.sheet_view.showGridLines = False

    def _create_data_sheet(
        self,
        wb: Workbook,
        df: pd.DataFrame,
        sheet_name: str,
        description: str = ""
    ):
        """Create a generic data sheet."""
        ws = wb.create_sheet(sheet_name[:31])

        ws.merge_cells('A1:F1')
        ws['A1'] = description or sheet_name
        ws['A1'].font = self.STYLES['title_font']
        ws['A1'].alignment = self.STYLES['center_align']
        ws.row_dimensions[1].height = 26

        headers = list(df.columns)
        self._write_header_row(ws, 3, headers)

        for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = (
                    '' if pd.isna(value)
                    else round(value, 4)
                    if isinstance(value, float)
                    else value
                )
                cell.font = self.STYLES['normal_font']
                if row_idx % 2 == 0:
                    cell.fill = self.STYLES['alt_row_fill']

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[
                get_column_letter(col_idx)
            ].width = 20

        ws.freeze_panes = 'A4'
        ws.sheet_view.showGridLines = False

    def _create_recon_sheet(self, wb: Workbook, findings: Dict):
        """Create reconciliation statistics sheet."""
        ws = wb.create_sheet("Reconciliation Stats")
        stats = findings.get('recon_stats', {})
        metrics = findings.get('variance_metrics', {})

        ws.merge_cells('A1:C1')
        ws['A1'] = "RECONCILIATION STATISTICS"
        ws['A1'].font = self.STYLES['title_font']
        ws['A1'].alignment = self.STYLES['center_align']

        self._write_header_row(ws, 3, ['Statistic', 'Value'])

        all_stats = {**stats, **metrics}
        for i, (key, value) in enumerate(all_stats.items(), start=4):
            ws[f'A{i}'] = key.replace('_', ' ').title()
            ws[f'B{i}'] = (
                round(value, 4)
                if isinstance(value, float)
                else str(value)
                if isinstance(value, list)
                else value
            )
            ws[f'A{i}'].font = self.STYLES['normal_font']
            ws[f'B{i}'].font = self.STYLES['bold_font']
            if i % 2 == 0:
                ws[f'A{i}'].fill = self.STYLES['alt_row_fill']
                ws[f'B{i}'].fill = self.STYLES['alt_row_fill']

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        ws.sheet_view.showGridLines = False

    def _create_multi_cover_sheet(
        self,
        wb: Workbook,
        multi_findings: Dict,
        report_name: str
    ):
        """Cover sheet for multi-file comparison report."""
        ws = wb.create_sheet("Cover Page")
        ws.merge_cells('A1:D1')
        ws['A1'] = "🔍 MULTI-FILE AUDIT COMPARISON REPORT"
        ws['A1'].font = Font(
            name='Calibri', bold=True, size=22, color='1F4E79'
        )
        ws['A1'].alignment = self.STYLES['center_align']
        ws.row_dimensions[1].height = 45

        info = [
            ("Report:", report_name.replace('_', ' ')),
            ("Generated:", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Files Compared:", multi_findings.get('file_count', 0)),
            ("Baseline File:", multi_findings.get('baseline_file', 'N/A')),
            ("Audit Engine:", "Python Auditor v2.0")
        ]

        for i, (label, value) in enumerate(info, start=3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = self.STYLES['bold_font']
            ws[f'B{i}'] = value
            ws[f'B{i}'].font = self.STYLES['normal_font']

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.sheet_view.showGridLines = False

    def _write_header_row(
        self,
        ws,
        row: int,
        headers: list
    ):
        """Write a styled header row to a worksheet."""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=str(header))
            cell.fill = self.STYLES['header_fill']
            cell.font = self.STYLES['header_font']
            cell.alignment = self.STYLES['center_align']
            cell.border = self.STYLES['thin_border']
        ws.row_dimensions[row].height = 24