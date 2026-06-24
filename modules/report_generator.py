"""
Report Generator - Fixed (no DataLabels import)
"""

import pandas as pd
import io
from datetime import datetime
from typing import Dict
import logging

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ReportGenerator:

    STYLES = {
        'header_fill': PatternFill('solid', fgColor='1F4E79'),
        'subheader_fill': PatternFill('solid', fgColor='2E75B6'),
        'alt_fill': PatternFill('solid', fgColor='DCE6F1'),
        'pass_fill': PatternFill('solid', fgColor='00B050'),
        'fail_fill': PatternFill('solid', fgColor='FF0000'),
        'high_fill': PatternFill('solid', fgColor='FF0000'),
        'medium_fill': PatternFill('solid', fgColor='FFC000'),
        'low_fill': PatternFill('solid', fgColor='00B0F0'),
        'header_font': Font(name='Calibri', bold=True, color='FFFFFF', size=11),
        'title_font': Font(name='Calibri', bold=True, color='1F4E79', size=14),
        'bold_font': Font(name='Calibri', bold=True, size=10),
        'normal_font': Font(name='Calibri', size=10),
        'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'left': Alignment(horizontal='left', vertical='center'),
        'thin_border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

    def _write_header(self, ws, row, headers, col_start=1):
        for i, h in enumerate(headers, col_start):
            c = ws.cell(row=row, column=i, value=str(h))
            c.fill = self.STYLES['header_fill']
            c.font = self.STYLES['header_font']
            c.alignment = self.STYLES['center']
            c.border = self.STYLES['thin_border']
        ws.row_dimensions[row].height = 22

    def _write_df(self, ws, df, start_row=1, start_col=1):
        if df.empty:
            return
        self._write_header(ws, start_row, list(df.columns), start_col)
        for ri, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
            for ci, val in enumerate(row, start=start_col):
                cell = ws.cell(row=ri, column=ci)
                if pd.isna(val):
                    cell.value = ''
                elif isinstance(val, float):
                    cell.value = round(val, 4)
                else:
                    cell.value = val
                cell.font = self.STYLES['normal_font']
                cell.border = self.STYLES['thin_border']
                if ri % 2 == 0:
                    cell.fill = self.STYLES['alt_fill']
        for ci, col in enumerate(df.columns, start=start_col):
            try:
                max_w = max(
                    len(str(col)),
                    df[col].astype(str).str.len().max()
                )
            except Exception:
                max_w = len(str(col))
            ws.column_dimensions[
                get_column_letter(ci)
            ].width = min(max_w + 4, 45)

    def generate_trial_balance_report(self, tb_df, entity_name, period):
        wb = Workbook()
        ws = wb.active
        ws.title = "Trial Balance"
        ws.merge_cells('A1:H1')
        ws['A1'] = f"TRIAL BALANCE — {entity_name.upper()}"
        ws['A1'].font = self.STYLES['title_font']
        ws['A1'].alignment = self.STYLES['center']
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Period: {period}"
        ws['A2'].alignment = self.STYLES['center']
        self._write_df(ws, tb_df, start_row=4)
        if not tb_df.empty and "Total Debits" in tb_df.columns:
            last_row = 4 + len(tb_df) + 1
            dr_col = list(tb_df.columns).index("Total Debits") + 1
            cr_col = list(tb_df.columns).index("Total Credits") + 1
            ws.cell(row=last_row, column=dr_col - 1, value="TOTALS").font = self.STYLES['bold_font']
            ws.cell(row=last_row, column=dr_col, value=tb_df["Total Debits"].sum()).font = self.STYLES['bold_font']
            ws.cell(row=last_row, column=cr_col, value=tb_df["Total Credits"].sum()).font = self.STYLES['bold_font']
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A5'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def generate_financial_report(self, is_data, bs_data, entity_name):
        wb = Workbook()
        wb.remove(wb.active)
        ws_is = wb.create_sheet("Income Statement")
        ws_is.merge_cells('A1:D1')
        ws_is['A1'] = f"INCOME STATEMENT — {entity_name.upper()}"
        ws_is['A1'].font = self.STYLES['title_font']
        ws_is['A1'].alignment = self.STYLES['center']
        row = 3
        sections = [
            ('REVENUE', is_data.get('revenue', pd.DataFrame())),
            ('COST OF SALES', is_data.get('cogs', pd.DataFrame())),
            ('OPERATING EXPENSES', is_data.get('expenses', pd.DataFrame()))
        ]
        for section_title, df in sections:
            cell = ws_is.cell(row=row, column=1, value=section_title)
            cell.fill = self.STYLES['subheader_fill']
            cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
            row += 1
            if not df.empty:
                for _, r in df.iterrows():
                    ws_is.cell(row=row, column=1, value=r.get('Account Code', '')).font = self.STYLES['normal_font']
                    ws_is.cell(row=row, column=2, value=r.get('Account Name', '')).font = self.STYLES['normal_font']
                    ws_is.cell(row=row, column=3, value=r.get('Closing Balance', 0)).font = self.STYLES['normal_font']
                    if row % 2 == 0:
                        for c in range(1, 4):
                            ws_is.cell(row=row, column=c).fill = self.STYLES['alt_fill']
                    row += 1
            row += 1
        totals = is_data.get('totals', {})
        summary_rows = [
            ('Gross Profit', totals.get('gross_profit', 0)),
            ('Operating Profit', totals.get('operating_profit', 0)),
            ('Net Income', totals.get('net_income', 0))
        ]
        for label, val in summary_rows:
            ws_is.cell(row=row, column=2, value=label).font = self.STYLES['bold_font']
            color = '00B050' if val >= 0 else 'FF0000'
            ws_is.cell(row=row, column=3, value=val).font = Font(
                name='Calibri', bold=True, color=color, size=11
            )
            row += 1
        ws_is.column_dimensions['A'].width = 16
        ws_is.column_dimensions['B'].width = 35
        ws_is.column_dimensions['C'].width = 18
        ws_is.sheet_view.showGridLines = False
        ws_bs = wb.create_sheet("Balance Sheet")
        ws_bs.merge_cells('A1:D1')
        ws_bs['A1'] = f"BALANCE SHEET — {entity_name.upper()}"
        ws_bs['A1'].font = self.STYLES['title_font']
        ws_bs['A1'].alignment = self.STYLES['center']
        assets = bs_data.get('assets', pd.DataFrame())
        if not assets.empty:
            drop_cols = [
                c for c in [
                    'Total Debits', 'Total Credits',
                    'Net Balance', 'Normal Balance'
                ] if c in assets.columns
            ]
            self._write_df(ws_bs, assets.drop(columns=drop_cols), start_row=3)
        ws_bs.sheet_view.showGridLines = False
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def generate_purchase_ledger_report(self, ledger_df, entity_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Ledger"
        ws.merge_cells('A1:I1')
        ws['A1'] = f"PURCHASE LEDGER — {entity_name.upper()}"
        ws['A1'].font = self.STYLES['title_font']
        ws['A1'].alignment = self.STYLES['center']
        self._write_df(ws, ledger_df, start_row=3)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = 'A4'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def generate_bin_card_report(self, bin_card_df, item_code, entity_name):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Bin Card {item_code}"
        ws.merge_cells('A1:J1')
        ws['A1'] = f"BIN CARD — {item_code} — {entity_name.upper()}"
        ws['A1'].font = self.STYLES['title_font']
        ws['A1'].alignment = self.STYLES['center']
        self._write_df(ws, bin_card_df, start_row=3)
        ws.sheet_view.showGridLines = False
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def generate_audit_report(self, findings, report_name="Audit_Report"):
        wb = Workbook()
        wb.remove(wb.active)
        ws_cover = wb.create_sheet("Cover")
        ws_cover.merge_cells('A1:D1')
        ws_cover['A1'] = f"AUDIT REPORT — {report_name}"
        ws_cover['A1'].font = self.STYLES['title_font']
        ws_cover['A1'].alignment = self.STYLES['center']
        stats = findings.get('recon_stats', {})
        metrics = findings.get('variance_metrics', {})
        info = [
            ("Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Comparison", findings.get('comparison_label', 'N/A')),
            ("Primary Key", findings.get('primary_key', 'N/A')),
            ("Match Rate", f"{stats.get('match_rate_pct', 0):.1f}%"),
            ("Total Differences", str(metrics.get('total_differences', 0))),
            ("Records Added", str(stats.get('records_added', 0))),
            ("Records Removed", str(stats.get('records_removed', 0))),
            ("Audit Status", "PASS" if stats.get('audit_pass') else "FAIL"),
        ]
        for i, (label, value) in enumerate(info, start=3):
            ws_cover.cell(row=i, column=1, value=label).font = self.STYLES['bold_font']
            ws_cover.cell(row=i, column=2, value=value).font = self.STYLES['normal_font']
        ws_cover.column_dimensions['A'].width = 25
        ws_cover.column_dimensions['B'].width = 35
        ws_cover.sheet_view.showGridLines = False
        flagged = findings.get('flagged_findings', pd.DataFrame())
        if not flagged.empty:
            ws_f = wb.create_sheet("Flagged Findings")
            self._write_df(ws_f, flagged, start_row=2)
            ws_f.sheet_view.showGridLines = False
        diff_df = findings.get('value_differences', pd.DataFrame())
        if not diff_df.empty:
            ws_d = wb.create_sheet("Value Changes")
            self._write_df(ws_d, diff_df, start_row=2)
            ws_d.sheet_view.showGridLines = False
        added = findings.get('added_records', pd.DataFrame())
        if not added.empty:
            ws_a = wb.create_sheet("Added Records")
            self._write_df(ws_a, added, start_row=2)
        removed = findings.get('removed_records', pd.DataFrame())
        if not removed.empty:
            ws_r = wb.create_sheet("Removed Records")
            self._write_df(ws_r, removed, start_row=2)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
