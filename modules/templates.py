"""
Upload Templates Module
Provides downloadable Excel templates for all data types
"""

import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class TemplateGenerator:
    """Generates Excel upload templates with instructions."""

    HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
    HEADER_FONT = Font(
        name='Calibri', bold=True, color='FFFFFF', size=11
    )
    EXAMPLE_FILL = PatternFill('solid', fgColor='E2EFDA')
    EXAMPLE_FONT = Font(name='Calibri', color='375623', size=10)
    CENTER = Alignment(horizontal='center', vertical='center')

    TEMPLATES = {
        'purchases': {
            'title': 'Purchase Transactions Upload Template',
            'columns': [
                'purchase_date', 'supplier_code', 'supplier_name',
                'invoice_number', 'invoice_date', 'due_date',
                'item_code', 'item_description', 'quantity',
                'unit_cost', 'tax_amount', 'discount_amount',
                'account_code', 'warehouse', 'notes'
            ],
            'example': [
                '2024-01-15', 'SUP001', 'ABC Suppliers Ltd',
                'INV-2024-001', '2024-01-15', '2024-02-14',
                'ITM001', 'Raw Material - Steel', 100,
                25.50, 0, 0, '1310', 'MAIN',
                'Monthly supply order'
            ],
            'instructions': [
                'Date format: YYYY-MM-DD',
                'Quantity: numeric (decimals allowed)',
                'Unit Cost: numeric, no currency symbols',
                'Account Code: from Chart of Accounts',
                'All amounts in company currency'
            ]
        },
        'purchase_returns': {
            'title': 'Purchase Returns Upload Template',
            'columns': [
                'return_date', 'original_purchase_number',
                'supplier_name', 'item_code', 'item_description',
                'quantity_returned', 'unit_cost', 'reason'
            ],
            'example': [
                '2024-01-20', 'PO-000001', 'ABC Suppliers Ltd',
                'ITM001', 'Raw Material - Steel (Defective)',
                10, 25.50, 'Defective goods received'
            ],
            'instructions': [
                'Return date must be after original purchase date',
                'Unit cost must match original purchase cost',
                'Reason is mandatory for audit trail'
            ]
        },
        'expenses': {
            'title': 'Expenses Upload Template',
            'columns': [
                'expense_date', 'expense_category', 'description',
                'vendor_name', 'reference', 'amount', 'tax_amount',
                'account_code', 'cost_center', 'payment_method',
                'notes'
            ],
            'example': [
                '2024-01-10', 'Utilities', 'Electricity Bill January',
                'City Power Corp', 'UTL-2024-001', 1250.00, 0,
                '6300', 'ADMIN', 'Bank Transfer', 'Monthly utility'
            ],
            'instructions': [
                'Category options: Utilities, Salaries, Rent, '
                'Travel, Marketing, Professional Fees, Other',
                'Payment Method: Cash, Bank Transfer, Credit Card',
                'Account Code from Chart of Accounts'
            ]
        },
        'journal_entries': {
            'title': 'Manual Journal Entry Upload Template',
            'columns': [
                'entry_date', 'entry_reference', 'description',
                'account_code', 'account_name',
                'debit_amount', 'credit_amount',
                'cost_center', 'notes'
            ],
            'example': [
                '2024-01-31', 'JE-ADJ-001',
                'Depreciation Adjustment January',
                '6400', 'Depreciation Expense',
                5000.00, 0, 'OPERATIONS', 'Monthly depreciation'
            ],
            'instructions': [
                'Each journal must balance: Total Debits = Total Credits',
                'Group lines by entry_reference',
                'Debit OR Credit per line (not both)',
                'Minimum 2 lines per journal entry'
            ]
        },
        'inventory_items': {
            'title': 'Inventory Master Upload Template',
            'columns': [
                'item_code', 'item_description', 'item_category',
                'unit_of_measure', 'reorder_level',
                'reorder_quantity', 'warehouse', 'bin_location'
            ],
            'example': [
                'ITM001', 'Raw Material - Steel Bars',
                'Raw Materials', 'KG', 50, 200,
                'MAIN', 'A-01-01'
            ],
            'instructions': [
                'Item Code must be unique',
                'UOM: KG, LITRE, UNIT, BOX, DOZEN, etc.',
                'Reorder Level: minimum stock before reorder',
                'Bin Location: physical storage location code'
            ]
        },
        'suppliers': {
            'title': 'Suppliers Upload Template',
            'columns': [
                'supplier_code', 'supplier_name', 'contact_person',
                'email', 'phone', 'address', 'payment_terms',
                'currency', 'tax_number', 'account_code', 'credit_limit'
            ],
            'example': [
                'SUP001', 'ABC Suppliers Ltd', 'John Smith',
                'john@abcsuppliers.com', '+1-555-0100',
                '123 Industrial Ave, New York, NY',
                30, 'USD', 'TAX123456', '2110', 50000
            ],
            'instructions': [
                'Supplier Code must be unique',
                'Payment Terms: days (e.g. 30 = Net 30)',
                'Credit Limit: maximum AP balance allowed'
            ]
        },
        'single_entry': {
            'title': 'Single Entry Transactions (for Conversion)',
            'columns': [
                'transaction_date', 'description', 'reference',
                'amount', 'transaction_type'
            ],
            'example': [
                '2024-01-15', 'Office supplies purchase', 'REC-001',
                250.00, 'EXPENSE_CASH'
            ],
            'instructions': [
                'Transaction Types: CASH_SALE, CREDIT_SALE, '
                'CASH_PURCHASE, CREDIT_PURCHASE, EXPENSE_CASH, '
                'EXPENSE_BANK, PAYMENT_SUPPLIER, RECEIPT_CUSTOMER, '
                'PAYROLL, DEPRECIATION, PURCHASE_RETURN',
                'Amount must be positive',
                'System will auto-generate the second leg'
            ]
        }
    }

    def generate_template(self, template_key: str) -> bytes:
        """Generate a styled Excel template for download."""
        if template_key not in self.TEMPLATES:
            raise ValueError(f"Unknown template: {template_key}")

        tmpl = self.TEMPLATES[template_key]
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Data Entry"

        # Title row
        last_col = get_column_letter(len(tmpl['columns']))
        ws_data.merge_cells(f'A1:{last_col}1')
        ws_data['A1'] = tmpl['title']
        ws_data['A1'].font = Font(
            name='Calibri', bold=True,
            color='1F4E79', size=14
        )
        ws_data['A1'].alignment = self.CENTER
        ws_data.row_dimensions[1].height = 30

        # Header row
        for ci, col in enumerate(tmpl['columns'], 1):
            c = ws_data.cell(row=2, column=ci, value=col.upper())
            c.fill = self.HEADER_FILL
            c.font = self.HEADER_FONT
            c.alignment = self.CENTER
            ws_data.row_dimensions[2].height = 20

        # Example row
        for ci, val in enumerate(tmpl['example'], 1):
            c = ws_data.cell(row=3, column=ci, value=val)
            c.fill = self.EXAMPLE_FILL
            c.font = self.EXAMPLE_FONT

        # Auto width
        for ci, col in enumerate(tmpl['columns'], 1):
            ws_data.column_dimensions[
                get_column_letter(ci)
            ].width = max(len(col) + 4, 18)

        ws_data.freeze_panes = 'A3'

        # Instructions sheet
        ws_inst = wb.create_sheet("Instructions")
        ws_inst['A1'] = "UPLOAD INSTRUCTIONS"
        ws_inst['A1'].font = Font(
            name='Calibri', bold=True,
            color='1F4E79', size=14
        )

        ws_inst['A2'] = "Template: " + tmpl['title']
        ws_inst['A2'].font = Font(
            name='Calibri', size=11, italic=True
        )

        ws_inst['A4'] = "Rules & Requirements:"
        ws_inst['A4'].font = Font(
            name='Calibri', bold=True, size=11
        )

        for ri, instruction in enumerate(
            tmpl['instructions'], start=5
        ):
            ws_inst[f'A{ri}'] = f"  • {instruction}"

        ws_inst.column_dimensions['A'].width = 80
        ws_inst.sheet_view.showGridLines = False

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def get_template_list(self):
        return [
            {
                'key': k,
                'title': v['title'],
                'columns': len(v['columns'])
            }
            for k, v in self.TEMPLATES.items()
        ]